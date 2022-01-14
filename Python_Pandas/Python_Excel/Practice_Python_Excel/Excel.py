from openpyxl import load_workbook
from data1 import datainfo1, datainfo2
from mapping import PRODUCT_ID, PRODUCT_PARENT, PRODUCT_TITLE, \
    PRODUCT_CATEGORY


path = '/Users/User/Desktop/Python/reviews-sample.xlsx'

workbook = load_workbook(filename= path, read_only=True)
sheet = workbook.active

products = []

# grabs a row two and all of its values
for row in sheet.iter_rows(min_row=2, values_only=True):
    product = datainfo1(id=row[PRODUCT_ID],
                      parent=row[PRODUCT_PARENT],
                      title=row[PRODUCT_TITLE],
                      category=row[PRODUCT_CATEGORY])
    products.append(product)

print("\n")
print(products[0])