from openpyxl import load_workbook

# workbook = load_workbook(filename="/Users/User/Desktop/Python/Data/reviews-sample.xlsx")
# workbook.sheetnames
# sheet = workbook.active

#______________________Table Contents_________________________________
# Printing entire excel sheet

# Print Row column

# Slicing through data 

# Iterating through an excel sheet 
    # printing rows
    # printing columns
    # print the values
    ## ALOT OF DATA ## Iterate through the whole dataset (rows and columns)

# Ouputting information into a dictionary

# Converting data into python Classes____________

# Sample Project extracting data from excel sheet
#_____________________End of Contents_____________________________


#__________________________________________________
# Printing entire excel sheet

# print(sheet.title)

# A1 = sheet["A1"]
# print(A1)

# A1 = sheet["A1"].value
# print(A1)

# F10 = sheet["F10"].value
# print(F10)


#__________________________________________________
# Row column
# test = sheet.cell(row=10, column=6)
# print(test)

# test = sheet.cell(row=10, column=6).value
# print(test)

#__________________________________________________
# Slicing through data 

# test = sheet["A1:C2"] # A1 - C2
# print(test)

# test = sheet["A"] # all of column A
# print(test)

# test = sheet["A:B"] # all of colmun A and B 
# print(test)

# test = sheet[5] # all of cells from row 5 
# print(test)

# test = sheet[5:6] # cells from A5 B5... N6 O6..ect  
# print(test)

# Iterating through an excel sheet 
#__________________________________________________
    ## Translation
        # min_row=1, max_row=2 : choosing two rows (1, 2)
        # min_col=1, max_col=3 : choosing columns (A, B, C) 

# printing the rows__________ 
# for row in sheet.iter_rows(min_row=1,
#     max_row=2,
#     min_col=1,
#     max_col=3):
#     print(row)

# printing the columns__________ 
# for column in sheet.iter_rows(min_row=1,
#     max_row=2,
#     min_col=1,
#     max_col=3):
#     print(column)

# print the values__________
# for value in sheet.iter_rows(min_row=1,
#     max_row=2,
#     min_col=1,
#     max_col=3,
#     values_only=True):  # values
#     print(value)

## ALOT OF DATA ## Iterate through the whole dataset (rows and columns) 
# for row in sheet.rows:
#     print(row)

# for column in sheet.columns:
#     print(column)

# test data__________
# grabing product_id, product_parent,  
#             product_title, product_category

# for value in sheet.iter_rows(min_row=1,
#     max_row=3,
#     min_col=4,
#     max_col=7,
#     values_only=True):
#     print(value)


# Ouputting information into a dictionary____________

# import json 

# products = {}

# for row in sheet.iter_rows(min_row=2,
#     min_col=4,
#     max_col=7,
#     values_only=True):

#     product_id = row[0]

#     product = {
#         "parent": row[1],
#         "title": row[2],
#         "category": row[3]
#     }
#     products[product_id] = product

# print(json.dumps(products))
##print(product) # prints the information readable format

#________________________________________________________________ 
# ________Sample Project extracting data from excel sheet________
#________________________________________________________________
# from datetime import datetime
# from openpyxl import load_workbook
# from classes import Product, Review
# from mapping import PRODUCT_ID, PRODUCT_PARENT, PRODUCT_TITLE, \
#     PRODUCT_CATEGORY, REVIEW_DATE, REVIEW_ID, REVIEW_CUSTOMER, \
#     REVIEW_STARS, REVIEW_HEADLINE, REVIEW_BODY

# # Using the read_only method since you're not gonna be editing the spreadsheet
# path = "/Users/User/Desktop/Python/Data/reviews-sample.xlsx"

# workbook = load_workbook(filename=path, read_only=True)
# sheet = workbook.active

# products = []
# reviews = []

# print("START HERE: ")

# # Using the values_only because you just want to return the cell value
# for row in sheet.iter_rows(min_row=2, values_only=True):
#     product = Product(id=row[PRODUCT_ID],
#                       parent=row[PRODUCT_PARENT],
#                       title=row[PRODUCT_TITLE],
#                       category=row[PRODUCT_CATEGORY])
#     products.append(product)

#     # You need to parse the date from the spreadsheet into a datetime format
#     spread_date = row[REVIEW_DATE]
#     parsed_date = datetime.strptime(spread_date, "%Y-%m-%d")

#     review = Review(id=row[REVIEW_ID],
#                     customer_id=row[REVIEW_CUSTOMER],
#                     stars=row[REVIEW_STARS],
#                     headline=row[REVIEW_HEADLINE],
#                     body=row[REVIEW_BODY],
#                     date=parsed_date)
#     reviews.append(review)

# print(products[0])
# print(reviews[0])