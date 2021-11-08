# Contents_______________________________________________________________
    # .value - printing the value of a cell  

# Contents_______________________________________________________________


from openpyxl import Workbook

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "Marketplace"
sheet["F10"] = "G-Shock Men's Grey Sport Watch"

workbook.save(filename="reviews-sample")

# prints value in F10 - To return the actual value of a cell, you need to do .value
print(sheet["F10"].value)
print(sheet.cell(row=10, column=6).value)


# slice through data [A1, B1, C1...]
print(sheet["A1:C2"])

# Get all columns from [A5 B5 C5.... A6 B6 C6]
print(sheet["A"])

print(sheet["A:B"])

# Get all cells from row 5 [A5 B5 C5...]
print(sheet[5])

# Get all cells range of rows [A5 B5 C5.... A6 B6 C6]
print(sheet[5:6])

for row in sheet.iter_rows(
    min_row=1,
    max_row=2,
    min_col=1,
    max_col=3,
    values_only=True):
    
    print(row)