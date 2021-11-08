#__________________Contents___________________
# Methods openpyxl

#________________End of Contents______________

from openpyxl import load_workbook

# Start by opening the spreadsheet and selecting the main sheet
workbook = load_workbook(filename="/Users/User/Desktop/Python/Python_Excel/hello_world.xlsx")
sheet = workbook.active

# Write what you want into a specific cell
sheet["A1"] = "hello"
sheet["B1"] = "world!"
# sheet["C1"] = "writing ;)"
# sheet["B10"] = "test"

# Save the spreadsheet
workbook.save(filename="/Users/User/Desktop/Python/Python_Excel/hello_world_append.xlsx")


# Useful function 
def print_rows():
    for row in sheet.iter_rows(values_only=True):
        print(row)

# Methods openpyxl______________ 
# .insert_rows()
# .delete_rows()
# .insert_cols()
# .delete_cols()

# Insert a column before the existing column 1 ("A")
sheet.insert_cols(idx=1)
print_rows()

# Insert 5 columns between column 2 ("B") and 3 ("C")
sheet.insert_cols(idx=3, amount=5)
print_rows()

# Delete the 5 previous created columns
sheet.delete_cols(idx=3, amount=5)
# sheet.delete_cols(idx=1)
print_rows()

# Insert a new row in the beginning
sheet.insert_rows(idx=1)
print_rows()

# Insert 3 new rows in the beginning
sheet.insert_rows(idx=1, amount=3)
print_rows()

# Delete the first 4 rows
sheet.delete_rows(idx=1, amount=4)
print_rows()

